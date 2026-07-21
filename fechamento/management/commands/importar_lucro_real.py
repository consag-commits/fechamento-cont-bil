"""
importar_lucro_real — Carrega a planilha "Clientes Lucro Real" na tela de
acompanhamento.

Espera uma aba com as colunas SEQUENCIA, EMPRESA, APURAÇÃO e SITUAÇÃO (o
cabeçalho é localizado automaticamente — a planilha original começa na linha 3).

A planilha chama as empresas pelo apelido ("UNIPRINT") e o cadastro guarda a
razão social inteira, então o casamento é feito por aproximação (ver _casar).
Empresa não encontrada é apenas listada: só é cadastrada com --criar-empresas,
para não duplicar quem já existe com o nome completo.

    uv run python manage.py importar_lucro_real planilha.xlsx --ano 2026 --simular
    uv run python manage.py importar_lucro_real planilha.xlsx --ano 2026
    uv run python manage.py importar_lucro_real planilha.xlsx --criar-empresas
"""

import unicodedata
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from fechamento.models import AcompanhamentoLucroReal, Empresa

_APURACAO = AcompanhamentoLucroReal.Apuracao

# Como o texto da planilha (já normalizado) vira uma opção do sistema.
_MAPA_APURACAO = {
    "MENSAL": _APURACAO.MENSAL,
    "TRIMESTRAL": _APURACAO.TRIMESTRAL,
    "RECEITA BRUTA": _APURACAO.RECEITA_BRUTA,
}


# Apelidos que não dá para deduzir do texto: a planilha abrevia de um jeito que
# não é prefixo nem subconjunto das palavras da razão social. Chave = como está
# na planilha; valor = razão social completa (a mesma do catálogo do Omie).
_APELIDOS = {
    "JL SANCHES IND E COMERCIO": "JL SANCHES INDUSTRIA E COMERCIO DE MATERIAIS PLASTICOS LTDA",
    "SENSE BIKE": "SENSE INDUSTRIA DE BICICLETAS DA AMAZONIA LTDA",
}


def _normalizar(texto):
    """Maiúsculas, sem acento e sem espaços sobrando — para comparar textos."""
    texto = str(texto or "").strip().upper()
    sem_acento = unicodedata.normalize("NFKD", texto)
    return " ".join("".join(c for c in sem_acento if not unicodedata.combining(c)).split())


def _razao_social(nome):
    """Razão social completa a procurar/cadastrar para um nome da planilha."""
    return _APELIDOS.get(_normalizar(nome), str(nome).strip())


class Command(BaseCommand):
    help = "Importa a planilha de Clientes Lucro Real para a tela de acompanhamento."

    def add_arguments(self, parser):
        parser.add_argument("planilha", help="Caminho do arquivo .xlsx.")
        parser.add_argument(
            "--ano", type=int, default=None,
            help="Ano do acompanhamento (padrão: ano atual).",
        )
        parser.add_argument(
            "--aba", default=None,
            help="Nome da aba a ler (padrão: a primeira).",
        )
        parser.add_argument(
            "--simular", action="store_true",
            help="Só mostra o que seria feito, sem gravar nada.",
        )
        parser.add_argument(
            "--criar-empresas", action="store_true",
            help="Cadastra as empresas que não forem encontradas. Sem esta opção "
                 "elas são apenas listadas — evita duplicar quem já existe com o "
                 "nome completo.",
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            raise CommandError("openpyxl não está instalado. Rode: uv sync --group dev")

        caminho = Path(options["planilha"])
        if not caminho.exists():
            raise CommandError(f"Planilha não encontrada: {caminho}")

        ano = options["ano"] or timezone.localdate().year

        wb = openpyxl.load_workbook(caminho, data_only=True)
        ws = wb[options["aba"]] if options["aba"] else wb[wb.sheetnames[0]]

        linhas = self._ler_linhas(ws)
        if not linhas:
            raise CommandError(
                "Não encontrei o cabeçalho (SEQUENCIA / EMPRESA / APURAÇÃO) na planilha."
            )

        simular = options["simular"]
        casamentos = self._casar(linhas)

        for nome, empresa, como in casamentos:
            if empresa is not None:
                self.stdout.write(f"  [{como}] {nome} -> #{empresa.pk} {empresa.razao_social}")
            elif como == "ambigua":
                self.stdout.write(self.style.WARNING(f"  [AMBÍGUA] {nome} — pulada, resolva na tela"))
            else:
                acao = "será criada" if options["criar_empresas"] else "PULADA (use --criar-empresas)"
                self.stdout.write(self.style.WARNING(f"  [SEM CADASTRO] {nome} — {acao}"))

        if simular:
            self.stdout.write(self.style.SUCCESS("\nSimulação — nada foi gravado."))
            return

        criadas, atualizadas, novas_empresas, puladas = 0, 0, 0, 0
        with transaction.atomic():
            for ordem, ((nome, apuracao_txt, situacao), (_n, empresa, _c)) in enumerate(
                zip(linhas, casamentos), start=1
            ):
                if empresa is None:
                    if not options["criar_empresas"]:
                        puladas += 1
                        continue
                    empresa = Empresa.objects.create(razao_social=_razao_social(nome))
                    novas_empresas += 1

                if not empresa.participa_lucro_real:
                    empresa.participa_lucro_real = True
                    empresa.save(update_fields=["participa_lucro_real"])

                apuracao = _MAPA_APURACAO.get(_normalizar(apuracao_txt), _APURACAO.NA)
                # "-" na planilha significa "sem observação".
                atualizacoes = "" if _normalizar(situacao) in ("", "-") else str(situacao).strip()

                _linha, criado = AcompanhamentoLucroReal.objects.update_or_create(
                    empresa=empresa, ano=ano,
                    defaults={"ordem": ordem, "apuracao": apuracao, "atualizacoes": atualizacoes},
                )
                criadas += criado
                atualizadas += not criado

        self.stdout.write(self.style.SUCCESS(
            f"\nLucro Real {ano}: {criadas} linha(s) criada(s), {atualizadas} atualizada(s)"
            f"{f', {novas_empresas} empresa(s) nova(s) cadastrada(s)' if novas_empresas else ''}"
            f"{f', {puladas} pulada(s) por não ter cadastro' if puladas else ''}."
        ))

    def _ler_linhas(self, ws):
        """[(empresa, apuração, situação)] a partir do cabeçalho encontrado."""
        colunas, dados = None, []
        for valores in ws.iter_rows(values_only=True):
            rotulos = [_normalizar(v) for v in valores]

            if colunas is None:
                if "EMPRESA" in rotulos:
                    colunas = {
                        "empresa": rotulos.index("EMPRESA"),
                        "apuracao": next((i for i, r in enumerate(rotulos) if r.startswith("APURACAO")), None),
                        "situacao": next(
                            (i for i, r in enumerate(rotulos) if r.startswith(("SITUACAO", "ATUALIZAC"))), None
                        ),
                    }
                continue

            nome = str(valores[colunas["empresa"]] or "").strip()
            if not nome:
                continue
            dados.append((
                nome,
                valores[colunas["apuracao"]] if colunas["apuracao"] is not None else "",
                valores[colunas["situacao"]] if colunas["situacao"] is not None else "",
            ))
        return dados

    def _casar(self, linhas):
        """[(nome, empresa|None, como)] — liga cada linha da planilha à empresa
        já cadastrada.

        A planilha usa apelidos curtos ("UNIPRINT") e o cadastro guarda a razão
        social inteira ("UNIPRINT COMERCIO DE SUPRIMENTOS DE INFORMATICA LTDA"),
        então vamos afrouxando o critério em três passadas — da mais segura para
        a mais solta. Cada empresa só pode ser usada por uma linha: é isso que
        desempata casos como STRONG SEGURANÇA / STRONG, em que o nome mais curto
        sobra para a empresa que a linha anterior não levou."""
        cadastro = list(Empresa.objects.all())
        normalizado = {e.pk: _normalizar(e.razao_social) for e in cadastro}
        usadas = set()
        resultado = [None] * len(linhas)

        def candidatas(alvo, criterio):
            return [
                e for e in cadastro
                if e.pk not in usadas and criterio(normalizado[e.pk], alvo)
            ]

        passadas = [
            ("exata", lambda razao, alvo: razao == alvo),
            ("prefixo", lambda razao, alvo: razao.startswith(alvo + " ")),
            ("palavras", lambda razao, alvo: all(p in razao.split() for p in alvo.split())),
        ]

        for como, criterio in passadas:
            for i, (nome, _apuracao, _situacao) in enumerate(linhas):
                if resultado[i] is not None:
                    continue
                achadas = candidatas(_normalizar(_razao_social(nome)), criterio)
                if len(achadas) == 1:
                    resultado[i] = (nome, achadas[0], como)
                    usadas.add(achadas[0].pk)

        for i, (nome, _apuracao, _situacao) in enumerate(linhas):
            if resultado[i] is None:
                # Sobrou ambígua (várias candidatas) ou sem nenhuma candidata.
                ambigua = len(candidatas(_normalizar(_razao_social(nome)), passadas[2][1])) > 1
                resultado[i] = (nome, None, "ambigua" if ambigua else "sem_cadastro")

        return resultado
